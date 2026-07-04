"""Liest die Header-Zeilen aller Blätter, um Spalten-Layouts zu vergleichen."""
from pathlib import Path

from openpyxl import load_workbook

xlsx = Path(__file__).resolve().parent.parent / "2026-06-05 Modelle.xlsx"
wb = load_workbook(xlsx, read_only=True, data_only=True)

for sheet in wb.worksheets:
    print(f"\n=== {sheet.title} ===")
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=3, values_only=True)):
        vals = [str(c)[:18] if c is not None else "·" for c in row[:10]]
        print(f"  Zeile {i+1}: {vals}")
wb.close()
