"""Prüft, ob die xlsx eingebettete Bilder ODER Bildpfade enthält."""
import zipfile
from pathlib import Path

xlsx = Path(__file__).resolve().parent.parent / "2026-06-05 Modelle.xlsx"

with zipfile.ZipFile(xlsx) as z:
    names = z.namelist()

media = [n for n in names if n.startswith("xl/media/")]
drawings = [n for n in names if "drawing" in n.lower()]

print("=== Eingebettete Bilder (xl/media/) ===")
if media:
    print(f"  JA — {len(media)} Bilder eingebettet")
    for m in media[:10]:
        print(f"    {m}")
else:
    print("  NEIN — keine eingebetteten Bilder")

print("=== Drawing-Referenzen ===")
print(f"  {len(drawings)} drawing-Dateien" if drawings else "  keine")

# Nach Bildpfad-Text in Zellen suchen (Hyperlinks / Dateipfade)
from openpyxl import load_workbook

wb = load_workbook(xlsx, read_only=True, data_only=True)
pfad_treffer = 0
beispiele = []
for sheet in wb.worksheets:
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and any(
                ext in cell.lower() for ext in (".jpg", ".jpeg", ".png", "http", "\\", "/bilder")
            ):
                pfad_treffer += 1
                if len(beispiele) < 5:
                    beispiele.append(cell[:80])
wb.close()

print("=== Bildpfade/URLs als Text in Zellen ===")
if pfad_treffer:
    print(f"  {pfad_treffer} verdächtige Zellen")
    for b in beispiele:
        print(f"    {b}")
else:
    print("  KEINE — keine Bildpfade/URLs in den Zellen")
