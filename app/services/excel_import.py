"""Excel-Import: die Sammler-Excel → SQLite.

HEADER-GETRIEBEN, weil die ~20 Blätter unterschiedliche Spalten-Layouts haben:
- Wiking Standard (UV500..W700): Nr|Min|Max|Model|Bemerkung|bezahlt|Schätzwert|Anzahl|Kaufdatum
- W800: ohne Anzahl-Spalte
- W600..W1000: 'Nr.' steht in Zeile 2 statt 1
- Siku/Matchbox/Majorette/Sonstige: Hersteller|Anzahl|Bezeichnung|Farbe|OVP|Zustand|Maßstab|Sonstiges|bezahlt|Schätzwert
- Siku: Siku(Nummer)|Anzahl|Bezeichnung|Farbe|OVP|Zustand|Sonstiges|Bezahlt|Kaufdatum
- playmobil: No.|Marke|Kaufdatum|Beschreibung|Anzahl|Total paid
- Summary/Wik Werbe: Sonderfälle

Strategie: Kopfzeilen (erste 2 Zeilen) je Spalte zusammenführen, per Keyword auf
kanonische Felder mappen, Datenzeilen danach lesen. Aggregat-/Summenblätter skippen.
"""
from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Katalog, Modell

# Blätter, die kein Modell-Inventar sind → überspringen
SKIP_SHEETS = {"summary"}

# Keyword → kanonisches Feld. Reihenfolge = Priorität (erste Übereinstimmung).
HEADER_MAP: list[tuple[str, str]] = [
    ("mindestwert", "min"), ("min.", "min"), ("wert katalog", "min"),
    ("höchstwert", "max"), ("max.", "max"),
    ("bemerkung", "bemerkung"), ("sonstiges", "sonstiges"),
    ("schätzwert", "schaetzwert"),
    ("bezahlt", "bezahlt"), ("total paid", "bezahlt"),
    ("anzahl", "anzahl"),
    ("kaufdatum", "kaufdatum"),
    ("farbe", "farbe"),
    ("ovp", "ovp"),
    ("zustand", "zustand_col"),
    ("maßstab", "massstab"),
    ("hersteller", "hersteller_col"), ("marke", "hersteller_col"),
    # Identität/Nummer + Typ zuletzt (breit)
    ("nummer", "nr"), ("bezeichnung", "typ"), ("beschreibung", "typ"),
    ("modell", "typ"), ("model", "typ"), ("typ", "typ"),
    ("nr.", "nr"), ("no.", "nr"), ("siku", "nr"),
]

_ZUSTAND_RE = re.compile(r"z\s*([012])", re.IGNORECASE)
_EXCEL_EPOCH = date(1899, 12, 30)


def _clean(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return round(float(v), 2)
    except (ValueError, TypeError):
        return None


def _to_int(v: Any, default: int = 1) -> int:
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def normalize_datum(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    try:
        serial = float(v)
        if 20000 < serial < 60000:
            return (_EXCEL_EPOCH + timedelta(days=int(serial))).isoformat()
        return None  # Zahl außerhalb Datumsbereich → kein Datum
    except (ValueError, TypeError):
        pass
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None  # unbekannt → lieber NULL als Müll


def extrahiere_zustand(*texte: str | None) -> str | None:
    """Findet z0/z1/z2 in beliebigen Textfeldern; Spanne → schlechtere Stufe."""
    ziffern: list[int] = []
    for t in texte:
        if t:
            ziffern += [int(x) for x in _ZUSTAND_RE.findall(t)]
    return f"z{max(ziffern)}" if ziffern else None


def _baue_header_map(rows: list[tuple]) -> tuple[dict[str, int], int]:
    """Aus den ersten Zeilen die Spalten→Feld-Zuordnung + Datenstart-Index bauen."""
    if not rows:
        return {}, 0
    breite = max(len(r) for r in rows)
    # Header können über Zeile 1+2 verteilt sein → pro Spalte zusammenführen
    header_zeilen = rows[:2]
    spalten_text: list[str] = []
    for col in range(breite):
        teile = []
        for hr in header_zeilen:
            if col < len(hr) and hr[col] is not None:
                teile.append(str(hr[col]).strip().lower())
        spalten_text.append(" ".join(teile))

    feld_zu_col: dict[str, int] = {}
    for col, text in enumerate(spalten_text):
        for keyword, feld in HEADER_MAP:
            if keyword in text and feld not in feld_zu_col:
                feld_zu_col[feld] = col
                break

    # Datenstart: erste Zeile nach den Headern, die in der nr/typ-Spalte einen
    # plausiblen Wert hat und nicht selbst Header ist.
    daten_start = 2
    id_col = feld_zu_col.get("nr", feld_zu_col.get("typ", 0))
    for i in range(1, min(4, len(rows))):
        val = rows[i][id_col] if id_col < len(rows[i]) else None
        sval = str(val).strip().lower() if val is not None else ""
        if sval and sval not in ("nr.", "nummer", "typ", "euro", "bezeichnung"):
            daten_start = i
            break
    return feld_zu_col, daten_start


def _ist_leer_oder_summe(zeile: tuple, fmap: dict[str, int]) -> bool:
    id_col = fmap.get("nr", fmap.get("typ", 0))
    typ_col = fmap.get("typ", id_col)
    idv = _clean(zeile[id_col]) if id_col < len(zeile) else None
    typv = _clean(zeile[typ_col]) if typ_col < len(zeile) else None
    if not idv and not typv:
        return True
    for marker in (idv, typv):
        if marker and marker.strip().lower() in ("nr.", "typ", "euro", "nummer", "gesamt"):
            return True
    return False


async def importiere_excel(
    session: AsyncSession, pfad: Path, default_hersteller: str = "Wiking"
) -> dict[str, int]:
    wb = load_workbook(pfad, read_only=True, data_only=True)
    stats = {"blaetter": 0, "katalog_neu": 0, "modelle": 0, "uebersprungen": 0, "skip_sheets": 0}

    result = await session.execute(
        select(Katalog.hersteller, Katalog.katalog_nr, Katalog.id)
    )
    katalog_cache: dict[tuple[str, str], int] = {
        (h, nr): kid for h, nr, kid in result.all()
    }

    def g(zeile: tuple, fmap: dict[str, int], feld: str) -> Any:
        col = fmap.get(feld)
        if col is None or col >= len(zeile):
            return None
        return zeile[col]

    for sheet in wb.worksheets:
        if sheet.title.strip().lower() in SKIP_SHEETS:
            stats["skip_sheets"] += 1
            continue
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 3:
            stats["skip_sheets"] += 1
            continue
        fmap, daten_start = _baue_header_map(rows)
        if "typ" not in fmap and "nr" not in fmap:
            stats["skip_sheets"] += 1
            continue

        serie = sheet.title
        # Hersteller: aus Spalte falls vorhanden, sonst je Blatt raten
        blatt_hersteller = default_hersteller
        low = serie.lower()
        if "siku" in low:
            blatt_hersteller = "Siku"
        elif "majorette" in low:
            blatt_hersteller = "Majorette"
        elif "matchbox" in low:
            blatt_hersteller = "Matchbox"
        elif "playmobil" in low:
            blatt_hersteller = "Playmobil"
        elif "sonstige" in low:
            blatt_hersteller = "Sonstige"

        stats["blaetter"] += 1

        for zeile in rows[daten_start:]:
            if not zeile or _ist_leer_oder_summe(zeile, fmap):
                stats["uebersprungen"] += 1
                continue
            typ = _clean(g(zeile, fmap, "typ")) or _clean(g(zeile, fmap, "nr"))
            if not typ:
                stats["uebersprungen"] += 1
                continue

            nr = _clean(g(zeile, fmap, "nr"))
            bemerkung = _clean(g(zeile, fmap, "bemerkung"))
            sonstiges = _clean(g(zeile, fmap, "sonstiges"))
            ovp = _clean(g(zeile, fmap, "ovp"))
            zustand_col = _clean(g(zeile, fmap, "zustand_col"))
            farbe = _clean(g(zeile, fmap, "farbe"))
            massstab = _clean(g(zeile, fmap, "massstab"))
            hersteller_col = _clean(g(zeile, fmap, "hersteller_col"))
            hersteller = hersteller_col or blatt_hersteller

            min_euro = _to_float(g(zeile, fmap, "min"))
            max_euro = _to_float(g(zeile, fmap, "max"))
            if bemerkung and "e.p." in bemerkung.lower():
                min_euro = min_euro or max_euro
                max_euro = max_euro or min_euro

            # Freitext-Bemerkung zusammensetzen (mehrere Quellspalten)
            bem_teile = [x for x in (bemerkung, zustand_col, ovp, sonstiges) if x]
            if massstab:
                bem_teile.append(f"Maßstab: {massstab}")
            voll_bemerkung = "; ".join(bem_teile) or None

            nr_key = nr or f"?/{serie}/{typ[:24]}"
            cache_key = (hersteller, nr_key)
            katalog_id = katalog_cache.get(cache_key)
            if katalog_id is None:
                kat = Katalog(
                    hersteller=hersteller,
                    katalog_nr=nr_key,
                    typ=typ,
                    serie=serie,
                    min_euro=min_euro,
                    max_euro=max_euro,
                    quelle="Rawe" if bemerkung and "rawe" in bemerkung.lower() else "GK",
                )
                session.add(kat)
                await session.flush()
                katalog_id = kat.id
                katalog_cache[cache_key] = katalog_id
                stats["katalog_neu"] += 1

            modell = Modell(
                katalog_id=katalog_id,
                farbe=farbe,
                zustand=extrahiere_zustand(zustand_col, bemerkung),
                bemerkung=voll_bemerkung,
                bezahlt=_to_float(g(zeile, fmap, "bezahlt")),
                schaetzwert=_to_float(g(zeile, fmap, "schaetzwert")),
                kaufdatum=normalize_datum(g(zeile, fmap, "kaufdatum")),
                anzahl=_to_int(g(zeile, fmap, "anzahl")),
                konvolut_id=None,
            )
            session.add(modell)
            stats["modelle"] += 1

    await session.commit()
    wb.close()
    return stats
