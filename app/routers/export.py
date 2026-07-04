"""Router: Excel-Import (Upload → SQLite) + Export."""
from __future__ import annotations

import io
import tempfile
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_session
from app.models import Katalog, Modell
from app.services.excel_import import importiere_excel

router = APIRouter(prefix="/api", tags=["import/export"])

SPALTEN = [
    "Nr.", "Min", "Max", "Typ", "Farbe", "Zustand",
    "Bemerkung", "bezahlt", "Schätzwert", "Anzahl", "Kaufdatum",
]


@router.post("/import/excel")
async def import_excel(
    datei: UploadFile,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Excel hochladen und in die DB importieren. Gibt Import-Statistik zurück."""
    name = (datei.filename or "").lower()
    if not name.endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Bitte eine .xlsx-Datei hochladen")

    inhalt = await datei.read()
    if not inhalt:
        raise HTTPException(400, "Leere Datei")
    if len(inhalt) > 25 * 1024 * 1024:  # 25 MB reicht für sehr große Sammlungen
        raise HTTPException(413, "Datei zu groß (max. 25 MB)")
    # openpyxl braucht einen Pfad/Stream — temporär auf Platte
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(inhalt)
        tmp_path = Path(tmp.name)
    try:
        stats = await importiere_excel(session, tmp_path)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(422, f"Import fehlgeschlagen: {e}") from e
    finally:
        tmp_path.unlink(missing_ok=True)

    return {"ok": True, "datei": datei.filename, **stats}


@router.get("/export/excel")
async def export_excel(session: AsyncSession = Depends(get_session)) -> StreamingResponse:
    stmt = select(Modell).join(Katalog).options(selectinload(Modell.katalog)).order_by(
        Katalog.hersteller, Katalog.katalog_nr
    )
    modelle = (await session.execute(stmt)).scalars().all()

    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)
    sheets: dict[str, Worksheet] = {}

    for m in modelle:
        h = m.katalog.hersteller or "Unbekannt"
        name = "".join(c for c in h if c not in "[]:*?/\\")[:31] or "Unbekannt"
        ws = sheets.get(name)
        if ws is None:
            ws = wb.create_sheet(name)
            ws.append(SPALTEN)
            sheets[name] = ws
        ws.append([
            m.katalog.katalog_nr, m.katalog.min_euro, m.katalog.max_euro,
            m.katalog.typ, m.farbe, m.zustand, m.bemerkung,
            m.bezahlt, m.schaetzwert, m.anzahl, m.kaufdatum,
        ])

    if not sheets:
        wb.create_sheet("Leer").append(SPALTEN)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    dateiname = f"ModellGarage_Export_{datetime.now():%Y-%m-%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{dateiname}"'},
    )
